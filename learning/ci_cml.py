import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from PyQt6.QtWidgets import (
    QApplication, QDialog, QFileDialog, QMessageBox, QWidget
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6 import uic



class CsvLoaderThread(QThread):
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    finished = pyqtSignal(pd.DataFrame)

    def __init__(self, file_path, chunksize=10000):
        super().__init__()
        self.file_path = file_path
        self.chunksize = chunksize

    def run(self):
        chunks = []
        try:
            with open(self.file_path, encoding='utf-8') as f:
                total_rows = sum(1 for _ in f) - 1

            if total_rows <= 0:
                self.status.emit("❌ CSV is empty.")
                self.finished.emit(pd.DataFrame())
                return

            loaded_rows = 0
            for chunk in pd.read_csv(self.file_path, chunksize=self.chunksize):
                chunks.append(chunk)
                loaded_rows += len(chunk)
                percent = int((loaded_rows / total_rows) * 100)
                self.progress.emit(percent)

            df = pd.concat(chunks)
            self.status.emit("✅ Done")
            self.finished.emit(df)

        except Exception as e:
            self.status.emit(f"❌ Error: {e}")
            self.finished.emit(pd.DataFrame())


class CICML(QDialog):
    def __init__(self, parent=None, flags=Qt.WindowType(0)):
        super().__init__(parent, flags)

        uic.loadUi('designs/self_things.ui', self)

        self.setWindowTitle('Test')
        self.resize(500, 250)

        self.submit_btn.setText('Submit')
        self.submit_btn.setFixedSize(70, 28)

        self.upload_btn.clicked.connect(self.upload_file)

        self.status_lbl.setText('Listening 👂')
        self.status_lbl.adjustSize()

        self.filename = ""
        self.file_path = ""

    def upload_file(self):
        self.file_path, _ = QFileDialog.getOpenFileName(
            self,
            caption='Upload file',
            directory="",
            filter="Supported Files: (*.csv *.xlsx *.xls *.xlsm)"
        )
        if self.file_path:
            self.filename = Path(self.file_path).name
            self.handle_file(self.file_path)


    def handle_file(self, filepath: str):
        ext = Path(filepath).suffix.lower()
        # self.filename = Path(filepath).name

        if ext == '.csv':
            self.load_csv(filepath)
        elif ext in {'.xls', '.xlsx', '.xlsm'}:
            self.load_excel(filepath)
        else:
            self.status_lbl.setText("❌ Unsupported file type.")


    def load_csv(self, file_path):
        self.thread = CsvLoaderThread(file_path)
        self.thread.progress.connect(self.update_progress)
        self.thread.status.connect(self.status_lbl.setText)
        self.thread.finished.connect(self.handle_csv_loaded)
        self.thread.start()

    
    def load_excel(self, file_path):
        dialog = SheetSelectorDialogue(file_path)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            sheetname = dialog.selected_sheet
        try:
            df = pd.read_excel(file_path, sheet_name=sheetname)
            if df.empty:
                self.status_lbl.setText("No data in Excel file.")
            else:
                self.status_lbl.setText(f"✅ Loaded {len(df)} rows from Excel.")
                self.main_status_lbl.setText(f"File: {self.filename}")
                self.setWindowTitle("Done")
        except Exception as e:
            self.status_lbl.setText(f"❌ Excel load error: {e}")



    def update_progress(self, percent):
        self.main_status_lbl.setText(f"Uploading: {self.filename}")
        self.status_lbl.setText(f"{percent}%")

    def handle_csv_loaded(self, df):
        if df.empty:
            self.status_lbl.setText("No data loaded.")
        else:
            self.status_lbl.setText(f"✅ Loaded {len(df)} rows.")
        self.setWindowTitle("Done")



class SheetSelectorDialogue(QDialog):
    def __init__(self, file_path, parent = None):
        super().__init__(parent)
        uic.loadUi(r'designs\sheet_selector.ui', self)

        self.setFixedSize(310, 240)

        self.setWindowTitle("Select Sheet")
        self.selected_sheet = None

        try:
            sheetnames = load_workbook(file_path, read_only=True).sheetnames
            self.sheet_name_cbox.addItems(['Select Sheet'] + sheetnames)
            self.sheet_name_cbox.currentTextChanged.connect(lambda text: self.sheet_name_cbox.setToolTip(text))

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read Excel file: {e}")
            self.reject()
        self.ok_btn.clicked.connect(self.accept_selection)
        self.cancel_btn.clicked.connect(self.reject)

    def accept_selection(self):
        selected = self.sheet_name_cbox.currentText()

        if selected == 'Select Sheet':
            QMessageBox.warning(self, 'Sheet Name Error', 'You forgot to select sheetname 🙂')
        else:
            self.selected_sheet = selected
            self.accept()




if __name__ == "__main__":
    # file = r'learning\LAD_DEC_2021_UK_NC.xlsx'
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = CICML()
    window.show()
    sys.exit(app.exec())
